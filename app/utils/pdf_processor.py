"""Processador de PDFs e Excel para anotações."""

import re
import uuid
from pathlib import Path
from typing import List

import fitz
import pdfplumber
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.note import Note
from app.models.user import User
from app.services.embedding_service import EmbeddingService


class PDFProcessor:
    """Processador de PDFs para converter em anotações."""

    @staticmethod
    def extract_structured_calendar_text(pdf_path: Path) -> str:
        """
        Extrai texto ESTRUTURADO de um PDF de calendário que está formatado como PLANILHA.
        
        Trata o PDF como uma planilha Excel, mantendo:
        - Estrutura de células e colunas
        - Alinhamento horizontal e vertical
        - Relação entre linhas e colunas
        - Organização por semanas e dias
        
        Args:
            pdf_path: Caminho para o arquivo PDF.
            
        Returns:
            str: Texto estruturado como planilha, bem organizado.
        """
        try:
            print(f"📅 Extraindo calendário estruturado (como PLANILHA) de: {pdf_path.name}")
            
            structured_parts = []
            
            # Usar pdfplumber para tratar como planilha
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"   📄 Processando página {page_num} como planilha...")
                    
                    # ESTRATÉGIA PRINCIPAL: Extrair como tabela/planilha
                    # Configurações otimizadas para planilhas
                    tables = page.extract_tables(
                        table_settings={
                            "vertical_strategy": "lines",  # Usar linhas verticais
                            "horizontal_strategy": "lines",  # Usar linhas horizontais
                            "explicit_vertical_lines": [],
                            "explicit_horizontal_lines": [],
                            "snap_tolerance": 5,  # Tolerância para alinhamento
                            "join_tolerance": 3,  # Tolerância para juntar células
                            "edge_tolerance": 3,  # Tolerância para bordas
                            "min_words_vertical": 1,  # Mínimo de palavras para coluna
                            "min_words_horizontal": 1,  # Mínimo de palavras para linha
                        }
                    )
                    
                    if tables:
                        print(f"      ✅ {len(tables)} tabela(s)/planilha(s) encontrada(s)")
                        for table_num, table in enumerate(tables, 1):
                            structured_parts.append(f"\n{'='*100}")
                            structured_parts.append(f"PLANILHA {table_num} - PÁGINA {page_num}")
                            structured_parts.append(f"{'='*100}\n")
                            
                            # Processar cada linha mantendo estrutura de colunas
                            for row_num, row in enumerate(table, 1):
                                if row:  # Linha não vazia
                                    # Processar células mantendo estrutura
                                    processed_cells = []
                                    for col_num, cell in enumerate(row, 1):
                                        if cell:
                                            cell_text = str(cell).strip()
                                            # Limpar quebras de linha e espaços múltiplos
                                            cell_text = re.sub(r'\s+', ' ', cell_text)
                                            # Manter célula mesmo se vazia (para manter estrutura)
                                            processed_cells.append(cell_text if cell_text else "")
                                        else:
                                            processed_cells.append("")  # Célula vazia
                                    
                                    # Juntar células com " | " para manter estrutura de colunas
                                    # Isso é CRÍTICO para a IA entender que são colunas diferentes
                                    row_text = " | ".join(processed_cells)
                                    
                                    # Adicionar número da linha para referência
                                    if row_text.strip():
                                        structured_parts.append(f"Linha {row_num:3d}: {row_text}")
                            
                            structured_parts.append("\n")
                    
                    # ESTRATÉGIA ALTERNATIVA: Se não encontrou tabelas, usar extração por palavras
                    # com agrupamento por posição (simula células de planilha)
                    if not tables or len(tables) == 0:
                        print(f"      📝 Extraindo como planilha por posição de palavras...")
                        
                        # Extrair palavras com posições precisas
                        words = page.extract_words(
                            x_tolerance=2,
                            y_tolerance=2,
                            keep_blank_chars=False,
                        )
                        
                        if words:
                            # Agrupar palavras por linha (Y similar) e depois por coluna (X similar)
                            lines_dict = {}
                            for word in words:
                                # Arredondar Y para agrupar linhas
                                y = round(word['top'] / 2) * 2
                                if y not in lines_dict:
                                    lines_dict[y] = []
                                lines_dict[y].append(word)
                            
                            # Processar cada linha
                            for y in sorted(lines_dict.keys()):
                                line_words = sorted(lines_dict[y], key=lambda w: w['x0'])
                                
                                # Agrupar palavras por coluna (X similar)
                                columns = []
                                current_col = []
                                prev_x = None
                                
                                for word in line_words:
                                    x = round(word['x0'] / 10) * 10  # Agrupar por X aproximado
                                    
                                    if prev_x is None or abs(x - prev_x) < 20:
                                        # Mesma coluna
                                        current_col.append(word['text'])
                                    else:
                                        # Nova coluna
                                        if current_col:
                                            columns.append(" ".join(current_col))
                                        current_col = [word['text']]
                                    
                                    prev_x = x
                                
                                if current_col:
                                    columns.append(" ".join(current_col))
                                
                                # Juntar colunas com " | " para manter estrutura
                                if columns:
                                    line_text = " | ".join(columns)
                                    if line_text.strip():
                                        structured_parts.append(line_text)
                    
                    # ESTRATÉGIA FALLBACK: Texto normal organizado
                    if not structured_parts or len(structured_parts) < 10:
                        text = page.extract_text()
                        if text and text.strip():
                            lines = text.split('\n')
                            for line in lines:
                                line = line.strip()
                                if line and len(line) > 2:
                                    structured_parts.append(line)
            
            # Juntar tudo
            full_text = "\n".join(structured_parts)
            
            # PÓS-PROCESSAMENTO: Organizar melhor o texto estruturado
            # 1. Identificar e marcar semanas (padrões comuns)
            full_text = re.sub(
                r'(?i)(semana\s*\d+|Semana\s*\d+|SEMANA\s*\d+)',
                r'\n\n' + '='*80 + '\n### \1 ###\n' + '='*80 + '\n',
                full_text
            )
            
            # 2. Identificar dias da semana (com destaque)
            full_text = re.sub(
                r'(?i)\b(Segunda|Terça|Quarta|Quinta|Sexta|Sábado|Domingo|Seg|Ter|Qua|Qui|Sex|Sáb|Dom)\b',
                r'\n--- \1 ---\n',
                full_text
            )
            
            # 3. Identificar plantões (com destaque especial)
            full_text = re.sub(
                r'(?i)(plantão|plantao|Plantão|PLANTÃO)',
                r'\n>>> 🚨 \1 🚨 <<<\n',
                full_text
            )
            
            # 4. Identificar preceptores
            full_text = re.sub(
                r'(?i)(preceptor|Preceptor|PRECEPTOR)',
                r'\n👨‍⚕️ \1',
                full_text
            )
            
            # 5. Limpar espaços extras mas manter estrutura
            full_text = re.sub(r'\n{4,}', '\n\n\n', full_text)  # Máximo 3 quebras
            full_text = re.sub(r' {3,}', ' ', full_text)  # Múltiplos espaços viram um
            full_text = re.sub(r'\n +', '\n', full_text)  # Espaços no início de linha
            
            # 6. Remover linhas completamente vazias duplicadas
            lines = full_text.split('\n')
            cleaned_lines = []
            prev_empty = False
            for line in lines:
                is_empty = not line.strip()
                if not (is_empty and prev_empty):
                    cleaned_lines.append(line)
                prev_empty = is_empty
            full_text = '\n'.join(cleaned_lines)
            
            print(f"✅ Calendário estruturado (planilha) extraído: {len(full_text)} caracteres")
            print(f"   📊 Estrutura: {full_text.count('PLANILHA')} planilhas, {full_text.count('Semana')} semanas, {full_text.count('|')} separadores de coluna")
            
            return full_text.strip()
            
        except Exception as e:
            print(f"❌ Erro ao extrair calendário estruturado: {e}")
            import traceback
            traceback.print_exc()
            # Fallback para extração simples
            print("⚠️ Usando extração simples como fallback...")
            return PDFProcessor.extract_text_from_pdf(pdf_path)
        """
        Extrai texto ESTRUTURADO de um PDF de calendário médico, organizando tabelas e dados.
        
        Esta função é específica para calendários médicos e:
        1. Extrai tabelas mantendo estrutura de colunas
        2. Organiza por semanas e dias da semana
        3. Identifica plantões, turnos e preceptores
        4. Mantém alinhamento e estrutura visual
        
        Args:
            pdf_path: Caminho para o arquivo PDF.
            
        Returns:
            str: Texto estruturado do calendário, bem organizado.
        """
        try:
            print(f"📅 Extraindo calendário estruturado de: {pdf_path.name}")
            
            structured_parts = []
            
            # Usar pdfplumber para extrair tabelas e texto estruturado
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"   📄 Processando página {page_num}...")
                    
                    # ESTRATÉGIA 1: Tentar extrair tabelas estruturadas
                    tables = page.extract_tables(
                        table_settings={
                            "vertical_strategy": "lines_strict",  # Usar linhas explícitas
                            "horizontal_strategy": "lines_strict",
                            "explicit_vertical_lines": [],
                            "explicit_horizontal_lines": [],
                            "snap_tolerance": 3,
                            "join_tolerance": 3,
                        }
                    )
                    
                    if tables:
                        print(f"      ✅ {len(tables)} tabela(s) encontrada(s)")
                        for table_num, table in enumerate(tables, 1):
                            structured_parts.append(f"\n{'='*80}")
                            structured_parts.append(f"TABELA {table_num} - PÁGINA {page_num}")
                            structured_parts.append(f"{'='*80}\n")
                            
                            # Processar cada linha da tabela mantendo estrutura
                            for row_num, row in enumerate(table, 1):
                                if row:  # Linha não vazia
                                    # Limpar e processar células
                                    clean_row = []
                                    for cell in row:
                                        if cell:
                                            cell_text = str(cell).strip()
                                            # Remover quebras de linha dentro da célula
                                            cell_text = re.sub(r'\s+', ' ', cell_text)
                                            if cell_text:
                                                clean_row.append(cell_text)
                                    
                                    if clean_row:
                                        # Juntar células com " | " para manter estrutura de colunas
                                        # Isso ajuda a IA a entender que são colunas diferentes
                                        row_text = " | ".join(clean_row)
                                        structured_parts.append(row_text)
                            
                            structured_parts.append("\n")  # Espaço entre tabelas
                    
                    # ESTRATÉGIA 2: Se não encontrou tabelas ou para complementar, usar extração por palavras
                    # Isso ajuda a manter ordem e posicionamento
                    if not tables or len(tables) == 0:
                        print(f"      📝 Extraindo texto estruturado por posição...")
                        
                        # Extrair palavras com posições
                        words = page.extract_words(
                            x_tolerance=3,
                            y_tolerance=3,
                        )
                        
                        if words:
                            # Agrupar palavras por linha (Y similar)
                            lines_dict = {}
                            for word in words:
                                y = round(word['top'] / 5) * 5  # Agrupar por Y aproximado
                                if y not in lines_dict:
                                    lines_dict[y] = []
                                lines_dict[y].append(word)
                            
                            # Ordenar linhas por Y (de cima para baixo)
                            for y in sorted(lines_dict.keys()):
                                line_words = sorted(lines_dict[y], key=lambda w: w['x0'])
                                line_text = " ".join([w['text'] for w in line_words])
                                if line_text.strip():
                                    structured_parts.append(line_text)
                    
                    # ESTRATÉGIA 3: Extrair texto normal como fallback
                    text = page.extract_text()
                    if text and text.strip():
                        # Se não encontrou tabelas, usar texto normal organizado
                        if not tables:
                            lines = text.split('\n')
                            for line in lines:
                                line = line.strip()
                                if line and len(line) > 2:  # Ignorar linhas muito curtas
                                    structured_parts.append(line)
            
            # Juntar tudo
            full_text = "\n".join(structured_parts)
            
            # PÓS-PROCESSAMENTO: Organizar melhor o texto
            # 1. Identificar e marcar semanas
            full_text = re.sub(
                r'(?i)(semana\s*\d+|Semana\s*\d+)',
                r'\n\n### \1 ###\n',
                full_text
            )
            
            # 2. Identificar dias da semana
            full_text = re.sub(
                r'(?i)\b(Segunda|Terça|Quarta|Quinta|Sexta|Sábado|Domingo|Seg|Ter|Qua|Qui|Sex|Sáb|Dom)\b',
                r'\n--- \1 ---\n',
                full_text
            )
            
            # 3. Identificar plantões (padrões comuns)
            full_text = re.sub(
                r'(?i)(plantão|plantao|Plantão)',
                r'\n>>> \1 <<<\n',
                full_text
            )
            
            # 4. Limpar espaços extras mas manter estrutura
            full_text = re.sub(r'\n{4,}', '\n\n\n', full_text)  # Máximo 3 quebras
            full_text = re.sub(r' {3,}', ' ', full_text)  # Múltiplos espaços viram um
            full_text = re.sub(r'\n +', '\n', full_text)  # Espaços no início de linha
            
            # 5. Remover linhas completamente vazias duplicadas
            lines = full_text.split('\n')
            cleaned_lines = []
            prev_empty = False
            for line in lines:
                is_empty = not line.strip()
                if not (is_empty and prev_empty):
                    cleaned_lines.append(line)
                prev_empty = is_empty
            full_text = '\n'.join(cleaned_lines)
            
            print(f"✅ Calendário estruturado extraído: {len(full_text)} caracteres")
            print(f"   📊 Estrutura: {full_text.count('TABELA')} tabelas, {full_text.count('Semana')} semanas identificadas")
            
            return full_text.strip()
            
        except Exception as e:
            print(f"❌ Erro ao extrair calendário estruturado: {e}")
            import traceback
            traceback.print_exc()
            # Fallback para extração simples
            print("⚠️ Usando extração simples como fallback...")
            return PDFProcessor.extract_text_from_pdf(pdf_path)

    @staticmethod
    def extract_structured_calendar_from_excel(excel_path: Path) -> str:
        """
        Extrai texto ESTRUTURADO de um arquivo Excel de calendário médico.
        
        Trata o Excel como planilha, mantendo:
        - Estrutura de células e colunas
        - Relação entre linhas e colunas
        - Organização por semanas e dias
        - Tabelas MAPA RECEPTOR e PLANTÃO
        
        Args:
            excel_path: Caminho para o arquivo Excel (.xlsx).
            
        Returns:
            str: Texto estruturado do calendário, bem organizado.
        """
        try:
            print(f"📊 Extraindo calendário estruturado de Excel: {excel_path.name}")
            
            structured_parts = []
            
            # Carregar workbook
            workbook = load_workbook(str(excel_path), data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"   📄 Processando planilha: {sheet_name}")
                
                structured_parts.append(f"\n{'='*100}")
                structured_parts.append(f"PLANILHA: {sheet_name}")
                structured_parts.append(f"{'='*100}\n")
                
                # Processar cada linha da planilha
                for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row:  # Linha não vazia
                        # Processar células mantendo estrutura
                        processed_cells = []
                        for col_num, cell in enumerate(row, 1):
                            if cell is not None:
                                cell_text = str(cell).strip()
                                # Limpar quebras de linha e espaços múltiplos
                                cell_text = re.sub(r'\s+', ' ', cell_text)
                                if cell_text:
                                    processed_cells.append(cell_text)
                                else:
                                    processed_cells.append("")  # Célula vazia
                            else:
                                processed_cells.append("")  # Célula None
                        
                        # Juntar células com " | " para manter estrutura de colunas
                        row_text = " | ".join(processed_cells)
                        
                        # Adicionar apenas se houver conteúdo
                        if row_text.strip() and row_text.strip() != "|":
                            structured_parts.append(f"Linha {row_num:3d}: {row_text}")
                
                structured_parts.append("\n")
            
            workbook.close()
            
            # Juntar tudo
            full_text = "\n".join(structured_parts)
            
            # PÓS-PROCESSAMENTO: Organizar melhor o texto estruturado
            # 1. Identificar e marcar semanas
            full_text = re.sub(
                r'(?i)(semana\s*\d+|Semana\s*\d+|SEMANA\s*\d+)',
                r'\n\n' + '='*80 + '\n### \1 ###\n' + '='*80 + '\n',
                full_text
            )
            
            # 2. Identificar tabelas MAPA RECEPTOR e PLANTÃO
            full_text = re.sub(
                r'(?i)(mapa\s*receptor|MAPA\s*RECEPTOR)',
                r'\n\n>>> MAPA RECEPTOR <<<\n',
                full_text
            )
            full_text = re.sub(
                r'(?i)(plantão|plantao|Plantão|PLANTÃO)',
                r'\n>>> 🚨 PLANTÃO 🚨 <<<\n',
                full_text
            )
            
            # 3. Identificar dias da semana
            full_text = re.sub(
                r'(?i)\b(Segunda|Terça|Quarta|Quinta|Sexta|Sábado|Domingo|Seg|Ter|Qua|Qui|Sex|Sáb|Dom)\b',
                r'\n--- \1 ---\n',
                full_text
            )
            
            # 4. Identificar preceptores
            full_text = re.sub(
                r'(?i)(preceptor|Preceptor|PRECEPTOR|principal\s*responsável)',
                r'\n👨‍⚕️ \1',
                full_text
            )
            
            # 5. Limpar espaços extras mas manter estrutura
            full_text = re.sub(r'\n{4,}', '\n\n\n', full_text)
            full_text = re.sub(r' {3,}', ' ', full_text)
            full_text = re.sub(r'\n +', '\n', full_text)
            
            # 6. Remover linhas completamente vazias duplicadas
            lines = full_text.split('\n')
            cleaned_lines = []
            prev_empty = False
            for line in lines:
                is_empty = not line.strip()
                if not (is_empty and prev_empty):
                    cleaned_lines.append(line)
                prev_empty = is_empty
            full_text = '\n'.join(cleaned_lines)
            
            print(f"✅ Calendário estruturado (Excel) extraído: {len(full_text)} caracteres")
            print(f"   📊 Estrutura: {full_text.count('PLANILHA')} planilhas, {full_text.count('Semana')} semanas, {full_text.count('|')} separadores de coluna")
            
            return full_text.strip()
            
        except Exception as e:
            print(f"❌ Erro ao extrair calendário do Excel: {e}")
            import traceback
            traceback.print_exc()
            raise ValueError(f"Erro ao processar arquivo Excel: {str(e)}")

    @staticmethod
    def extract_text_from_pdf(pdf_path: Path) -> str:
        """
        Extrai texto SIMPLES de um PDF apenas para embeddings (RAG).
        
        IMPORTANTE: Não tenta formatar ou estruturar. Apenas extrai texto puro.

        Args:
            pdf_path: Caminho para o arquivo PDF.

        Returns:
            str: Texto extraído do PDF (sem formatação).
        """
        try:
            # Extração SIMPLES apenas para RAG/embeddings
            print(f"📄 Extraindo texto de: {pdf_path.name}")
            
            text_parts = []
            
            # Usar PyMuPDF (fitz) para extração simples
            doc = fitz.open(str(pdf_path))
            
            for page_num, page in enumerate(doc, 1):
                # Extrair texto simples da página
                text = page.get_text("text")
                if text.strip():
                    text_parts.append(text.strip())
            
            doc.close()
            
            # Juntar todo o texto
            full_text = "\n\n".join(text_parts)
            
            # Limpar espaços extras
            full_text = re.sub(r'\n{3,}', '\n\n', full_text)
            full_text = re.sub(r' +', ' ', full_text)
            
            print(f"✅ Extraídos {len(full_text)} caracteres de {pdf_path.name}")
            
            return full_text.strip()
            
        except Exception as e:
            print(f"❌ Erro ao extrair PDF: {e}")
            return ""

    @staticmethod
    def chunk_text(text: str, chunk_size: int = 3000, overlap: int = 200) -> List[str]:
        """
        Divide texto em chunks com overlap.

        Args:
            text: Texto para dividir.
            chunk_size: Tamanho máximo de cada chunk.
            overlap: Overlap entre chunks.

        Returns:
            List[str]: Lista de chunks de texto.
        """
        if len(text) <= chunk_size:
            return [text]

        chunks = []
        start = 0

        while start < len(text):
            end = start + chunk_size

            # Se não é o último chunk, tenta encontrar um ponto de quebra natural
            if end < len(text):
                # Tenta quebrar em parágrafo
                last_newline = text.rfind("\n\n", start, end)
                if last_newline != -1 and last_newline > start + chunk_size // 2:
                    end = last_newline
                else:
                    # Tenta quebrar em sentença
                    last_period = text.rfind(". ", start, end)
                    if last_period != -1 and last_period > start + chunk_size // 2:
                        end = last_period + 1

            chunks.append(text[start:end].strip())
            start = end - overlap if end < len(text) else end

        return chunks

    @staticmethod
    async def process_pdf_to_notes(
        pdf_path: Path,
        user: User,
        db: AsyncSession,
        tag: str = "documento_importado",
        auto_index: bool = True,
    ) -> List[Note]:
        """
        Processa um PDF e cria anotações no banco de dados.

        Args:
            pdf_path: Caminho para o arquivo PDF.
            user: Usuário dono das anotações.
            db: Sessão do banco de dados.
            tag: Tag para as anotações criadas.
            auto_index: Se deve indexar automaticamente (gerar embeddings).

        Returns:
            List[Note]: Lista de anotações criadas.
        """
        # Extrair texto do PDF
        text = PDFProcessor.extract_text_from_pdf(pdf_path)

        # Dividir em chunks
        chunks = PDFProcessor.chunk_text(text)

        # Criar anotações
        notes = []
        for i, chunk in enumerate(chunks, 1):
            # Título baseado no nome do arquivo
            title = f"{pdf_path.stem} - Parte {i}/{len(chunks)}"

            # Criar nota
            note = Note(
                user_id=user.id,
                title=title,
                content=chunk,
                tags=[tag, "pdf", pdf_path.stem.lower().replace(" ", "_")],
                is_favorite=False,
            )

            db.add(note)
            notes.append(note)

        # Commit de todas as notas
        await db.commit()

        # Indexar (gerar embeddings)
        if auto_index:
            for note in notes:
                await db.refresh(note)
                try:
                    await EmbeddingService.create_or_update_embedding(note, db)
                except Exception as e:
                    print(f"Erro ao indexar nota {note.id}: {e}")

        return notes

    @staticmethod
    async def process_pdf_as_document(
        pdf_path: Path,
        user: User,
        db: AsyncSession,
        description: str = "",
    ):
        """
        Processa um PDF e cria um Document (apenas para RAG, não cria anotação).
        
        Args:
            pdf_path: Caminho para o arquivo PDF.
            user: Usuário dono do documento.
            db: Sessão do banco de dados.
            description: Descrição opcional do documento.
            
        Returns:
            Document: Documento criado ou None se houve erro.
        """
        from app.models.document import Document
        from app.models.document_embedding import DocumentEmbedding
        
        try:
            print(f"📄 Processando PDF como documento: {pdf_path.name}")
            
            # Extrair texto SIMPLES para embeddings
            text = PDFProcessor.extract_text_from_pdf(pdf_path)
            
            if not text:
                print(f"❌ Não foi possível extrair texto de: {pdf_path.name}")
                return None
            
            # Obter tamanho do arquivo
            file_size = pdf_path.stat().st_size
            
            # Criar documento
            document = Document(
                user_id=user.id,
                filename=pdf_path.name,
                file_path=str(pdf_path),
                file_size=file_size,
                description=description or f"Documento PDF: {pdf_path.stem}",
            )
            
            db.add(document)
            await db.commit()
            await db.refresh(document)
            
            print(f"✅ Documento criado: {document.id}")
            
            # Criar embedding (limitar tamanho para não exceder limite da API)
            try:
                print(f"🔄 Gerando embedding...")
                
                # Limite de 30.000 caracteres (~30KB) para ser seguro
                # Google Gemini tem limite de 36.000 bytes
                text_for_embedding = text[:30000] if len(text) > 30000 else text
                
                if len(text) > 30000:
                    print(f"⚠️ Texto muito grande ({len(text)} chars). Usando primeiros 30.000 caracteres.")
                
                embedding_vector = EmbeddingService.generate_embedding(text_for_embedding)
                
                document_embedding = DocumentEmbedding(
                    document_id=document.id,
                    content_preview=text[:500],  # Primeiros 500 chars para preview
                    embedding=embedding_vector,
                )
                
                db.add(document_embedding)
                await db.commit()
                
                print(f"✅ Embedding criado para documento: {document.id}")
                
            except Exception as e:
                print(f"❌ Erro ao criar embedding: {e}")
                # Documento foi criado mesmo sem embedding
            
            return document
            
        except Exception as e:
            print(f"❌ Erro ao processar PDF: {e}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def process_directory(
        directory_path: Path,
        user: User,
        db: AsyncSession,
        tag: str = "documento_importado",
        pattern: str = "*.pdf",
    ) -> dict:
        """
        Processa todos os PDFs de um diretório.

        Args:
            directory_path: Caminho para o diretório.
            user: Usuário dono das anotações.
            db: Sessão do banco de dados.
            tag: Tag para as anotações criadas.
            pattern: Padrão de arquivos (glob).

        Returns:
            dict: Estatísticas do processamento.
        """
        pdf_files = list(directory_path.glob(pattern))

        total_files = len(pdf_files)
        total_notes = 0
        errors = []

        for pdf_file in pdf_files:
            try:
                notes = await PDFProcessor.process_pdf_to_notes(
                    pdf_path=pdf_file,
                    user=user,
                    db=db,
                    tag=tag,
                    auto_index=True,
                )
                total_notes += len(notes)
                print(f"✅ Processado: {pdf_file.name} ({len(notes)} anotações)")
            except Exception as e:
                errors.append({"file": pdf_file.name, "error": str(e)})
                print(f"❌ Erro ao processar {pdf_file.name}: {e}")

        return {
            "total_files": total_files,
            "total_notes": total_notes,
            "errors": errors,
            "success_rate": (total_files - len(errors)) / total_files if total_files > 0 else 0,
        }

